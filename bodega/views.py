from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from django.utils.crypto import get_random_string
from django.core.mail import send_mail
from django.conf import settings
from django.urls import reverse, reverse_lazy
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView
from django.views import View
from django.contrib.auth.views import PasswordResetConfirmView
from django.utils.decorators import method_decorator
from django.forms import inlineformset_factory
from datetime import datetime
import openpyxl
from django.utils.timezone import localtime
from .models import Movimiento
from .decorators import permiso_requerido
from openpyxl import Workbook
from django.db.models import Q
from django.utils.encoding import smart_str
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
from num2words import num2words

from .models import ( 
    Empleado, Usuario, Rol, Permiso, Producto, Movimiento, Proveedor, OrdenCompra, DetalleOrdenCompra, Cliente,
    Almacen, CategoriaProducto, Inventario, Pedido, DetallePedido, Factura, DetalleFactura, Caja, MovimientoCaja,
    TipoProducto
)
from .forms import (
    EmpleadoForm, UsuarioForm, RolForm, PermisoForm, CambiarPasswordForm, ProductoForm, MovimientoForm, ProveedorForm,
    OrdenCompraForm, DetalleOrdenCompraForm, ClienteForm, AlmacenForm, CategoriaProductoForm, PedidoForm, DetallePedidoFormSet,
    FacturaForm, DetalleFacturaForm, DetalleOrdenCompraFormSet, CajaForm, MovimientoCajaForm,
    TipoProductoForm,
) 
from django.contrib.auth.tokens import default_token_generator
from .decorators import permiso_requerido
from django.forms import modelformset_factory
from django.shortcuts import get_object_or_404


DetalleFormSet = inlineformset_factory(
    OrdenCompra,
    DetalleOrdenCompra,
    form=DetalleOrdenCompraForm,
    extra=3,
    can_delete=False
)

DetalleFacturaFormSet = inlineformset_factory(
    Factura,
    DetalleFactura,
    form=DetalleFacturaForm,
    extra=3,
    can_delete=False
)

# Vista Home
@login_required
def home(request):
    return render(request, 'bodega/home.html')

# --- Empleados ---
@permiso_requerido('crear_empleado')
def crear_empleado(request):
    if request.method == 'POST':
        form = EmpleadoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home')
    else:
        form = EmpleadoForm()
    return render(request, 'bodega/empleado_form.html', {'form': form})


@method_decorator(permiso_requerido('ver_empleado'), name='dispatch')
class EmpleadoListView(ListView):
    model = Empleado
    template_name = "bodega/empleado_list.html"
    context_object_name = "empleados"
    paginate_by = 20

    def get_queryset(self):
        qs = Empleado.objects.select_related("sucursal")  # evita N+1 en sucursal
        q = (self.request.GET.get("q") or "").strip()
        estado = (self.request.GET.get("estado") or "").strip()

        if q:
            qs = qs.filter(
                Q(nombre__icontains=q)
                | Q(apellido__icontains=q)
                | Q(cedula__icontains=q)
                | Q(email__icontains=q)
                | Q(telefono__icontains=q)
            )

        if estado == "activos":
            qs = qs.filter(activo=True)
        elif estado == "inactivos":
            qs = qs.filter(activo=False)

        return qs.order_by("-activo", "apellido", "nombre", "id")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = (self.request.GET.get("q") or "").strip()
        ctx["estado"] = (self.request.GET.get("estado") or "").strip()
        ctx["total"] = Empleado.objects.count()
        # cantidad luego de aplicar filtros (si hay paginator, úsalo)
        ctx["filtrados"] = getattr(
            ctx.get("paginator"), "count", len(ctx.get("empleados", []))
        )
        return ctx


# class EmpleadoListView(LoginRequiredMixin, ListView):
#     model = Empleado
#     template_name = 'bodega/empleado_list.html'
#     context_object_name = 'empleados'
#     paginate_by = 10

#     def get_queryset(self):
#         queryset = Empleado.objects.all()
#         busqueda = self.request.GET.get('buscar')
#         if busqueda:
#             queryset = queryset.filter(nombre__icontains=busqueda) | queryset.filter(apellido__icontains=busqueda)
#         return queryset

@method_decorator(permiso_requerido('editar_empleado'), name='dispatch')
class EmpleadoUpdateView(LoginRequiredMixin, UpdateView):
    model = Empleado
    form_class = EmpleadoForm
    template_name = 'bodega/empleado_form.html'
    success_url = reverse_lazy('empleado_list')

@method_decorator(permiso_requerido('inactivar_empleado'), name='dispatch')
class EmpleadoInactivateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        self.object = Empleado.objects.get(pk=kwargs['pk'])
        return render(request, 'bodega/empleado_confirm_inactivate.html', {'object': self.object})

    def post(self, request, *args, **kwargs):
        self.object = Empleado.objects.get(pk=kwargs['pk'])
        self.object.activo = not self.object.activo
        self.object.save()
        return redirect('empleado_list')

# --- Usuarios ---
@method_decorator(permiso_requerido('crear_usuario'), name='dispatch')
class UsuarioCreateView(LoginRequiredMixin, CreateView):
    model = Usuario
    form_class = UsuarioForm
    template_name = 'bodega/usuario_form.html'
    success_url = reverse_lazy('usuario_list')

    def form_valid(self, form):
        usuario = form.save(commit=False)
        temp_password = get_random_string(length=8)
        usuario.password = make_password(temp_password)
        usuario.estado = 'activo'
        usuario.is_active = False
        usuario.save()
        form.save_m2m()

        # token = token_generator.make_token(usuario)
        token = default_token_generator.make_token(usuario)
        uid = urlsafe_base64_encode(force_bytes(usuario.pk))
        reset_url = self.request.build_absolute_uri(
            reverse('cambiar_password', kwargs={'uidb64': uid, 'token': token})
        )

        subject = 'Bienvenido a Bodega Eirete - Cambia tu contraseña'
        message = (
            f"Hola {usuario.username},\n\n"
            f"Se ha creado tu cuenta en Bodega Eirete.\n\n"
            f"Tu contraseña temporal es: {temp_password}\n\n"
            f"Por seguridad, ingresa en el siguiente enlace para cambiar tu contraseña:\n{reset_url}\n\n"
            f"Gracias."
        )
        send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [usuario.email])
        return HttpResponseRedirect(self.success_url)

@method_decorator(permiso_requerido('ver_usuario'), name='dispatch')
class UsuarioListView(LoginRequiredMixin, ListView):
    model = Usuario
    template_name = 'bodega/usuario_list.html'
    context_object_name = 'usuarios'
    paginate_by = 10

    def get_queryset(self):
        queryset = Usuario.objects.select_related('empleado', 'rol')
        busqueda = self.request.GET.get('buscar')
        if busqueda:
            queryset = queryset.filter(
                username__icontains=busqueda
            ) | queryset.filter(email__icontains=busqueda)
        return queryset

@method_decorator(permiso_requerido('editar_usuario'), name='dispatch')
class UsuarioUpdateView(LoginRequiredMixin, UpdateView):
    model = Usuario
    form_class = UsuarioForm
    template_name = 'bodega/usuario_form.html'
    success_url = reverse_lazy('usuario_list')

@method_decorator(permiso_requerido('inactivar_usuario'), name='dispatch')
class UsuarioInactivateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        self.object = Usuario.objects.get(pk=kwargs['pk'])
        return render(request, 'bodega/usuario_confirm_inactivate.html', {'object': self.object})

    def post(self, request, *args, **kwargs):
        self.object = Usuario.objects.get(pk=kwargs['pk'])
        self.object.estado = 'activo' if self.object.estado == 'inactivo' else 'inactivo'
        self.object.save()
        return redirect('usuario_list')

# --- Roles ---
@method_decorator(permiso_requerido('ver_rol'), name='dispatch')
class RolListView(LoginRequiredMixin, ListView):
    model = Rol
    template_name = 'bodega/rol_list.html'
    context_object_name = 'roles'
    paginate_by = 10

    def get_queryset(self):
        queryset = Rol.objects.all()
        busqueda = self.request.GET.get('buscar')
        if busqueda:
            queryset = queryset.filter(
                nombre__icontains=busqueda
            ) | queryset.filter(descripcion__icontains=busqueda)
        return queryset

@method_decorator(permiso_requerido('crear_rol'), name='dispatch')
class RolCreateView(LoginRequiredMixin, CreateView):
    model = Rol
    form_class = RolForm
    template_name = 'bodega/rol_form.html'
    success_url = reverse_lazy('rol_list')

@method_decorator(permiso_requerido('editar_rol'), name='dispatch')
class RolUpdateView(LoginRequiredMixin, UpdateView):
    model = Rol
    form_class = RolForm
    template_name = 'bodega/rol_form.html'
    success_url = reverse_lazy('rol_list')

@method_decorator(permiso_requerido('inactivar_rol'), name='dispatch')
class RolInactivateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        self.object = Rol.objects.get(pk=kwargs['pk'])
        return render(request, 'bodega/rol_confirm_inactivate.html', {'object': self.object})

    def post(self, request, *args, **kwargs):
        self.object = Rol.objects.get(pk=kwargs['pk'])
        self.object.activo = not self.object.activo
        self.object.save()
        return redirect('rol_list')

@method_decorator(permiso_requerido('editar_rol'), name='dispatch')
class RolAsignarPermisosView(LoginRequiredMixin, UpdateView):
    model = Rol
    form_class = RolForm
    template_name = 'bodega/rol_form.html'
    success_url = reverse_lazy('rol_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        self.object.permisos.set(form.cleaned_data['permisos'])
        return response

# --- Permisos ---
@method_decorator(permiso_requerido('ver_permiso'), name='dispatch')
class PermisoListView(LoginRequiredMixin, ListView):
    model = Permiso
    template_name = 'bodega/permiso_list.html'
    context_object_name = 'permisos'
    paginate_by = 10

    def get_queryset(self):
        queryset = Permiso.objects.all().order_by('id')
        busqueda = self.request.GET.get('buscar')
        if busqueda:
            queryset = queryset.filter(
                nombre__icontains=busqueda
            ) | queryset.filter(descripcion__icontains=busqueda)
        return queryset

@method_decorator(permiso_requerido('crear_permiso'), name='dispatch')
class PermisoCreateView(LoginRequiredMixin, CreateView):
    model = Permiso
    form_class = PermisoForm
    template_name = 'bodega/permiso_form.html'
    success_url = reverse_lazy('permiso_list')

@method_decorator(permiso_requerido('editar_permiso'), name='dispatch')
class PermisoUpdateView(LoginRequiredMixin, UpdateView):
    model = Permiso
    form_class = PermisoForm
    template_name = 'bodega/permiso_form.html'
    success_url = reverse_lazy('permiso_list')

@method_decorator(permiso_requerido('inactivar_permiso'), name='dispatch')
class PermisoInactivateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        self.object = Permiso.objects.get(pk=kwargs['pk'])
        return render(request, 'bodega/permiso_confirm_inactivate.html', {'object': self.object})

    def post(self, request, *args, **kwargs):
        self.object = Permiso.objects.get(pk=kwargs['pk'])
        self.object.activo = not self.object.activo
        self.object.save()
        return redirect('permiso_list')

class PrimerCambioPasswordView(PasswordResetConfirmView):
    template_name = 'bodega/cambiar_password.html'
    success_url = reverse_lazy('login')
    form_class = CambiarPasswordForm
    token_generator = default_token_generator

    def form_valid(self, form):
        response = super().form_valid(form)
        self.user.is_active = True
        self.user.save()
        return response

@method_decorator(permiso_requerido('ver_producto'), name='dispatch')
class ProductoListView(LoginRequiredMixin, ListView):
    model = Producto
    template_name = 'bodega/producto_list.html'
    context_object_name = 'productos'
    paginate_by = 10

    def get_queryset(self):
        queryset = Producto.objects.all().order_by('nombre')
        busqueda = self.request.GET.get('buscar')
        if busqueda:
            queryset = queryset.filter(
                nombre__icontains=busqueda
            ) | queryset.filter(descripcion__icontains=busqueda)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        return context


@method_decorator(permiso_requerido('crear_producto'), name='dispatch')
class ProductoCreateView(LoginRequiredMixin, CreateView):
    model = Producto
    form_class = ProductoForm
    template_name = 'bodega/producto_form.html'
    success_url = reverse_lazy('producto_list')

@method_decorator(permiso_requerido('editar_producto'), name='dispatch')
class ProductoUpdateView(LoginRequiredMixin, UpdateView):
    model = Producto
    form_class = ProductoForm
    template_name = 'bodega/producto_form.html'
    success_url = reverse_lazy('producto_list')

@method_decorator(permiso_requerido('inactivar_producto'), name='dispatch')
class ProductoInactivateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        self.object = Producto.objects.get(pk=kwargs['pk'])
        return render(request, 'bodega/producto_confirm_inactivate.html', {'object': self.object})

    def post(self, request, *args, **kwargs):
        self.object = Producto.objects.get(pk=kwargs['pk'])
        self.object.activo = not self.object.activo
        self.object.save()
        return redirect('producto_list')

@method_decorator(permiso_requerido('ver_movimiento'), name='dispatch')
class MovimientoListView(LoginRequiredMixin, ListView):
    model = Movimiento
    template_name = 'bodega/movimiento_list.html'
    context_object_name = 'movimientos'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().select_related('producto', 'realizado_por').order_by('-fecha')
        buscar = self.request.GET.get("buscar")
        producto_id = self.request.GET.get("producto_id")

        if buscar:
            queryset = queryset.filter(producto__nombre__icontains=buscar)

        if producto_id:
            queryset = queryset.filter(producto__id=producto_id)

        return queryset

@method_decorator(permiso_requerido('crear_movimiento'), name='dispatch')
class MovimientoCreateView(LoginRequiredMixin, CreateView):
    model = Movimiento
    form_class = MovimientoForm
    template_name = 'bodega/movimiento_form.html'
    success_url = reverse_lazy('movimiento_list')

    def form_valid(self, form):
        movimiento = form.save(commit=False)
        movimiento.realizado_por = self.request.user
        producto = movimiento.producto
        almacen = movimiento.almacen

        # Obtener inventario en el almacén
        inventario, _ = Inventario.objects.get_or_create(
            producto=producto,
            almacen=almacen
        )

        stock_actual = inventario.stock

        # Validar stock disponible en caso de salida
        if movimiento.tipo == 'salida' and movimiento.cantidad > stock_actual:
            form.add_error('cantidad', f'La cantidad supera el stock disponible en el almacén seleccionado ({stock_actual}).')
            return self.form_invalid(form)

        # Actualizar stock del producto total
        if movimiento.tipo == 'entrada':
            producto.stock += movimiento.cantidad
            inventario.stock += movimiento.cantidad
        elif movimiento.tipo == 'salida':
            producto.stock -= movimiento.cantidad
            inventario.stock -= movimiento.cantidad

        producto.save()
        inventario.save()
        movimiento.save()

        return redirect(self.success_url)

@method_decorator(permiso_requerido('ver_proveedor'), name='dispatch')
class ProveedorListView(LoginRequiredMixin, ListView):
    model = Proveedor
    template_name = 'bodega/proveedor_list.html'
    context_object_name = 'proveedores'
    paginate_by = 10

    def get_queryset(self):
        queryset = Proveedor.objects.all().order_by('nombre')
        buscar = self.request.GET.get('buscar')
        if buscar:
            queryset = queryset.filter(nombre__icontains=buscar)
        return queryset

@method_decorator(permiso_requerido('crear_proveedor'), name='dispatch')
class ProveedorCreateView(LoginRequiredMixin, CreateView):
    model = Proveedor
    form_class = ProveedorForm
    template_name = 'bodega/proveedor_form.html'
    success_url = reverse_lazy('proveedor_list')

@method_decorator(permiso_requerido('editar_proveedor'), name='dispatch')
class ProveedorUpdateView(LoginRequiredMixin, UpdateView):
    model = Proveedor
    form_class = ProveedorForm
    template_name = 'bodega/proveedor_form.html'
    success_url = reverse_lazy('proveedor_list')

@method_decorator(permiso_requerido('inactivar_proveedor'), name='dispatch')
class ProveedorInactivateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        self.object = Proveedor.objects.get(pk=kwargs['pk'])
        return render(request, 'bodega/proveedor_confirm_inactivate.html', {'object': self.object})

    def post(self, request, *args, **kwargs):
        self.object = Proveedor.objects.get(pk=kwargs['pk'])
        self.object.activo = not self.object.activo
        self.object.save()
        return redirect('proveedor_list')

@method_decorator(permiso_requerido('ver_orden_compra'), name='dispatch')
class OrdenCompraListView(LoginRequiredMixin, ListView):
    model = OrdenCompra
    template_name = 'bodega/orden_compra_list.html'
    context_object_name = 'ordenes'
    paginate_by = 10

    def get_queryset(self):
        qs = OrdenCompra.objects.select_related('proveedor').order_by('-id')
        buscar = self.request.GET.get('buscar')
        if buscar:
            qs = qs.filter(proveedor__nombre__icontains=buscar)
        return qs

@method_decorator(permiso_requerido('crear_orden_compra'), name='dispatch')
class OrdenCompraCreateView(LoginRequiredMixin, View):
    def get(self, request):
        orden_form = OrdenCompraForm()
        formset = DetalleFormSet()
        return render(request, 'bodega/orden_compra_form.html', {
            'orden_form': orden_form,
            'formset': formset
        })
    def post(self, request, pk=None):
        instance = None
        if pk:
            instance = get_object_or_404(OrdenCompra, pk=pk)

        orden_form = OrdenCompraForm(request.POST, instance=instance)
        formset = DetalleOrdenCompraFormSet(request.POST, instance=instance)

        if orden_form.is_valid() and formset.is_valid():
            orden = orden_form.save()
            formset.instance = orden
            formset.save()
            return redirect('orden_compra_list')

        return render(request, 'ordencompra_form.html', {
            'orden_form': orden_form,
            'formset': formset,
            'object': instance
    })


    # def post(self, request):
    #     orden_form = OrdenCompraForm(request.POST)
    #     formset = DetalleFormSet(request.POST)

    #     if orden_form.is_valid() and formset.is_valid():
    #         orden = orden_form.save()
    #         formset.instance = orden  # CORRECCIÓN
    #         formset.save()
    #         return redirect('orden_compra_list')

    #     return render(request, 'bodega/orden_compra_form.html', {
    #         'orden_form': orden_form,
    #         'formset': formset
    #     })

# @method_decorator(permiso_requerido('editar_orden_compra'), name='dispatch')
# class OrdenCompraUpdateView(LoginRequiredMixin, View):
#     def get(self, request, pk):
#         orden = OrdenCompra.objects.get(pk=pk)
#         orden_form = OrdenCompraForm(instance=orden)
#         formset = DetalleFormSet(instance=orden)
#         return render(request, 'bodega/orden_compra_form.html', {
#             'orden_form': orden_form,
#             'formset': formset
#         })

#     def post(self, request, pk):
#         orden = OrdenCompra.objects.get(pk=pk)
#         orden_form = OrdenCompraForm(request.POST, instance=orden)
#         formset = DetalleFormSet(request.POST, instance=orden)

#         if orden_form.is_valid() and formset.is_valid():
#             orden = orden_form.save()
#             formset.instance = orden  # CORRECCIÓN
#             formset.save()
#             return redirect('orden_compra_list')

#         return render(request, 'bodega/orden_compra_form.html', {
#             'orden_form': orden_form,
#             'formset': formset
#         })

DetalleOrdenCompraFormSet = inlineformset_factory(
    OrdenCompra,
    DetalleOrdenCompra,
    form=DetalleOrdenCompraForm,
    extra=1,
    can_delete=True
)

@method_decorator(permiso_requerido('editar_orden_compra'), name='dispatch')
class OrdenCompraUpdateView(LoginRequiredMixin, View):
    def get(self, request, pk):
        orden = get_object_or_404(OrdenCompra, pk=pk)
        orden_form = OrdenCompraForm(instance=orden)
        formset = DetalleOrdenCompraFormSet(instance=orden)
        return render(request, 'bodega/orden_compra_form.html', {
            'orden_form': orden_form,
            'formset': formset,
            'object': orden
        })

    def post(self, request, pk):
        orden = get_object_or_404(OrdenCompra, pk=pk)
        orden_form = OrdenCompraForm(request.POST, instance=orden)
        formset = DetalleOrdenCompraFormSet(request.POST, instance=orden)

        if orden_form.is_valid() and formset.is_valid():
            orden = orden_form.save()
            formset.instance = orden
            formset.save()  # ← ESTA LÍNEA AHORA BORRA CORRECTAMENTE LOS DETALLES MARCADOS
            return redirect('orden_compra_list')

        return render(request, 'bodega/orden_compra_form.html', {
            'orden_form': orden_form,
            'formset': formset,
            'object': orden
        })

# class OrdenCompraUpdateView(LoginRequiredMixin, View):
#     def get(self, request, pk):
#         orden = get_object_or_404(OrdenCompra, pk=pk)
#         orden_form = OrdenCompraForm(instance=orden)
#         formset = DetalleOrdenCompraFormSet(instance=orden)
#         return render(request, 'bodega/orden_compra_form.html', {
#             'orden_form': orden_form,
#             'formset': formset,
#             'object': orden
#         })

#     def post(self, request, pk):
#         orden = get_object_or_404(OrdenCompra, pk=pk)
#         orden_form = OrdenCompraForm(request.POST, instance=orden)
#         formset = DetalleOrdenCompraFormSet(request.POST, instance=orden)

#         if orden_form.is_valid() and formset.is_valid():
#             orden_form.save()
#             formset.save()
#             return redirect('orden_compra_list')
#         else:
#             return render(request, 'bodega/orden_compra_form.html', {
#                 'orden_form': orden_form,
#                 'formset': formset,
#                 'object': orden
#             })

@method_decorator(permiso_requerido('recibir_orden_compra'), name='dispatch')
class OrdenCompraRecibirView(LoginRequiredMixin, View):
    def get(self, request, pk):
        orden = OrdenCompra.objects.get(pk=pk)
        if orden.estado != 'pendiente':
            return redirect('orden_compra_list')  # Evita recibir dos veces

        return render(request, 'bodega/orden_compra_confirm_recibir.html', {
            'orden': orden
        })

    def post(self, request, pk):
        orden = OrdenCompra.objects.get(pk=pk)
        if orden.estado == 'pendiente':
            if not orden.almacen_destino:
                # Protección: no procesar si no hay almacén definido
                from django.contrib import messages
                messages.error(request, "La orden no tiene definido un almacén destino.")
                return redirect('orden_compra_list')

            for detalle in orden.detalles.all():
                producto = detalle.producto
                producto.stock += detalle.cantidad
                producto.save()

                # Actualizar inventario del almacén destino
                inventario, _ = Inventario.objects.get_or_create(
                    producto=producto,
                    almacen=orden.almacen_destino
                )
                inventario.stock += detalle.cantidad
                inventario.save()

            orden.estado = 'recibido'
            orden.save()

        return redirect('orden_compra_list')


# @method_decorator(permiso_requerido('recibir_orden_compra'), name='dispatch')
# class OrdenCompraRecibirView(LoginRequiredMixin, View):
#     def get(self, request, pk):
#         orden = OrdenCompra.objects.get(pk=pk)
#         if orden.estado != 'pendiente':
#             return redirect('orden_compra_list')  # Evita recibir dos veces

#         return render(request, 'bodega/orden_compra_confirm_recibir.html', {
#             'orden': orden
#         })

#     def post(self, request, pk):
#         orden = OrdenCompra.objects.get(pk=pk)
#         if orden.estado == 'pendiente':
#             for detalle in orden.detalles.all():
#                 producto = detalle.producto
#                 producto.stock += detalle.cantidad
#                 producto.save()

#                 # Actualizar inventario por almacén
#                 inventario, _ = Inventario.objects.get_or_create(
#                     producto=producto,
#                     almacen=orden.almacen_destino
#                 )
#                 inventario.stock += detalle.cantidad
#                 inventario.save()
#             orden.estado = 'recibido'
#             orden.save()
#         return redirect('orden_compra_list')

@method_decorator(permiso_requerido('cancelar_orden_compra'), name='dispatch')
class OrdenCompraCancelarView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        orden = OrdenCompra.objects.get(pk=kwargs['pk'])
        return render(request, 'bodega/orden_compra_confirm_cancelar.html', {'orden': orden})

    def post(self, request, *args, **kwargs):
        orden = OrdenCompra.objects.get(pk=kwargs['pk'])
        if orden.estado == 'pendiente':
            orden.estado = 'cancelado'
            orden.save()
        return redirect('orden_compra_list')

@method_decorator(permiso_requerido('ver_cliente'), name='dispatch')
class ClienteListView(LoginRequiredMixin, ListView):
    model = Cliente
    template_name = 'bodega/cliente_list.html'
    context_object_name = 'clientes'
    paginate_by = 10

    def get_queryset(self):
        queryset = Cliente.objects.all()
        buscar = self.request.GET.get('buscar')
        if buscar:
            queryset = queryset.filter(nombre__icontains=buscar) | queryset.filter(apellido__icontains=buscar)
        return queryset

@method_decorator(permiso_requerido('crear_cliente'), name='dispatch')
class ClienteCreateView(LoginRequiredMixin, CreateView):
    model = Cliente
    form_class = ClienteForm
    template_name = 'bodega/cliente_form.html'
    success_url = reverse_lazy('cliente_list')

@method_decorator(permiso_requerido('editar_cliente'), name='dispatch')
class ClienteUpdateView(LoginRequiredMixin, UpdateView):
    model = Cliente
    form_class = ClienteForm
    template_name = 'bodega/cliente_form.html'
    success_url = reverse_lazy('cliente_list')

@method_decorator(permiso_requerido('inactivar_cliente'), name='dispatch')
class ClienteInactivateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        self.object = Cliente.objects.get(pk=kwargs['pk'])
        return render(request, 'bodega/cliente_confirm_inactivate.html', {'object': self.object})

    def post(self, request, *args, **kwargs):
        self.object = Cliente.objects.get(pk=kwargs['pk'])
        self.object.activo = not self.object.activo
        self.object.save()
        return redirect('cliente_list')

# views.py
@method_decorator(permiso_requerido('ver_almacen'), name='dispatch')
class AlmacenListView(LoginRequiredMixin, ListView):
    model = Almacen
    template_name = 'bodega/almacen_list.html'
    context_object_name = 'almacenes'
    paginate_by = 10

    def get_queryset(self):
        queryset = Almacen.objects.all().order_by('nombre')
        busqueda = self.request.GET.get("buscar")
        if busqueda:
            queryset = queryset.filter(nombre__icontains=busqueda)
        return queryset


@method_decorator(permiso_requerido('crear_almacen'), name='dispatch')
class AlmacenCreateView(LoginRequiredMixin, CreateView):
    model = Almacen
    form_class = AlmacenForm
    template_name = 'bodega/almacen_form.html'
    success_url = reverse_lazy('almacen_list')

@method_decorator(permiso_requerido('editar_almacen'), name='dispatch')
class AlmacenUpdateView(LoginRequiredMixin, UpdateView):
    model = Almacen
    form_class = AlmacenForm
    template_name = 'bodega/almacen_form.html'
    success_url = reverse_lazy('almacen_list')

@method_decorator(permiso_requerido('inactivar_almacen'), name='dispatch')
class AlmacenInactivateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        self.object = Almacen.objects.get(pk=kwargs['pk'])
        return render(request, 'bodega/almacen_confirm_inactivate.html', {'object': self.object})

    def post(self, request, *args, **kwargs):
        self.object = Almacen.objects.get(pk=kwargs['pk'])
        self.object.activo = not self.object.activo
        self.object.save()
        return redirect('almacen_list')

@method_decorator(permiso_requerido('ver_categoria'), name='dispatch')
class CategoriaProductoListView(LoginRequiredMixin, ListView):
    model = CategoriaProducto
    template_name = 'bodega/categoria_list.html'
    context_object_name = 'categorias'
    paginate_by = 10

    def get_queryset(self):
        queryset = CategoriaProducto.objects.all().order_by('nombre')
        buscar = self.request.GET.get('buscar')
        if buscar:
            queryset = queryset.filter(nombre__icontains=buscar)
        return queryset

@method_decorator(permiso_requerido('crear_categoria'), name='dispatch')
class CategoriaProductoCreateView(LoginRequiredMixin, CreateView):
    model = CategoriaProducto
    form_class = CategoriaProductoForm
    template_name = 'bodega/categoria_form.html'
    success_url = reverse_lazy('categoria_list')

@method_decorator(permiso_requerido('editar_categoria'), name='dispatch')
class CategoriaProductoUpdateView(LoginRequiredMixin, UpdateView):
    model = CategoriaProducto
    form_class = CategoriaProductoForm
    template_name = 'bodega/categoria_form.html'
    success_url = reverse_lazy('categoria_list')

@method_decorator(permiso_requerido('inactivar_categoria'), name='dispatch')
class CategoriaProductoInactivateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        self.object = CategoriaProducto.objects.get(pk=kwargs['pk'])
        return render(request, 'bodega/categoria_confirm_inactivate.html', {'object': self.object})

    def post(self, request, *args, **kwargs):
        self.object = CategoriaProducto.objects.get(pk=kwargs['pk'])
        self.object.activo = not self.object.activo
        self.object.save()
        return redirect('categoria_list')

@method_decorator(permiso_requerido('ver_inventario'), name='dispatch')
class InventarioListView(LoginRequiredMixin, ListView):
    model = Inventario
    template_name = 'bodega/inventario_list.html'
    context_object_name = 'inventarios'
    paginate_by = 10

    def get_queryset(self):
        # qs = Inventario.objects.select_related('producto', 'almacen')
        qs = Inventario.objects.select_related('producto', 'producto__categoria', 'almacen')
        buscar = self.request.GET.get('buscar')
        if buscar:
            qs = qs.filter(
                producto__nombre__icontains=buscar
            ) | qs.filter(
                almacen__nombre__icontains=buscar
            )
        return qs.order_by('producto__nombre')

@method_decorator(permiso_requerido('ver_pedido'), name='dispatch')
class PedidoListView(LoginRequiredMixin, ListView):
    model = Pedido
    template_name = 'bodega/pedido_list.html'
    context_object_name = 'pedidos'
    paginate_by = 10

    def get_queryset(self):
        queryset = Pedido.objects.select_related('cliente').order_by('-fecha')
        buscar = self.request.GET.get('buscar')
        if buscar:
            queryset = queryset.filter(cliente__nombre__icontains=buscar) | queryset.filter(cliente__apellido__icontains=buscar)
        return queryset

@method_decorator(permiso_requerido('crear_pedido'), name='dispatch')
class PedidoCreateView(LoginRequiredMixin, View):
    def get(self, request):
        pedido_form = PedidoForm()
        formset = DetallePedidoFormSet(queryset=DetallePedido.objects.none())
        return render(request, 'bodega/pedido_form.html', {
            'form': pedido_form,
            'formset': formset
        })

    def post(self, request):
        pedido_form = PedidoForm(request.POST)
        formset = DetallePedidoFormSet(request.POST)
        if pedido_form.is_valid() and formset.is_valid():
            pedido = pedido_form.save()
            detalles = formset.save(commit=False)
            for detalle in detalles:
                detalle.pedido = pedido
                detalle.save()
            return redirect('pedido_list')
        return render(request, 'bodega/pedido_form.html', {
            'form': pedido_form,
            'formset': formset
        })

@method_decorator(permiso_requerido('editar_pedido'), name='dispatch')
class PedidoUpdateView(LoginRequiredMixin, UpdateView):
    model = Pedido
    form_class = PedidoForm
    template_name = 'bodega/pedido_form.html'

    def get(self, request, pk):
        pedido = get_object_or_404(Pedido, pk=pk)
        form = PedidoForm(instance=pedido)
        formset = DetallePedidoFormSet(instance=pedido)
        return render(request, self.template_name, {
            'form': form,
            'formset': formset,
            'object': pedido
        })

    def post(self, request, pk):
        pedido = get_object_or_404(Pedido, pk=pk)
        form = PedidoForm(request.POST, instance=pedido)
        formset = DetallePedidoFormSet(request.POST, instance=pedido)

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect('pedido_list')

        return render(request, self.template_name, {
            'form': form,
            'formset': formset,
            'object': pedido
        })

# @method_decorator(permiso_requerido('editar_pedido'), name='dispatch')
# class PedidoUpdateView(LoginRequiredMixin, View):
#     def get(self, request, pk):
#         pedido = Pedido.objects.get(pk=pk)
#         pedido_form = PedidoForm(instance=pedido)
#         formset = DetallePedidoFormSet(instance=pedido)
#         return render(request, 'bodega/pedido_form.html', {
#             'form': pedido_form,
#             'formset': formset
#         })

#     def post(self, request, pk):
#         pedido = Pedido.objects.get(pk=pk)
#         pedido_form = PedidoForm(request.POST, instance=pedido)
#         formset = DetallePedidoFormSet(request.POST, instance=pedido)
#         if pedido_form.is_valid() and formset.is_valid():
#             pedido_form.save()
#             formset.save()
#             return redirect('pedido_list')
#         return render(request, 'bodega/pedido_form.html', {
#             'form': pedido_form,
#             'formset': formset
#         })

@method_decorator(permiso_requerido('cancelar_pedido'), name='dispatch')
class PedidoInactivateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        pedido = Pedido.objects.get(pk=kwargs['pk'])
        return render(request, 'bodega/pedido_confirm_inactivate.html', {'pedido': pedido})

    def post(self, request, *args, **kwargs):
        pedido = Pedido.objects.get(pk=kwargs['pk'])
        if pedido.estado != 'cancelado':
            pedido.estado = 'cancelado'
        else:
            pedido.estado = 'pendiente'
        pedido.save()
        return redirect('pedido_list')

@method_decorator(permiso_requerido('editar_detallepedido'), name='dispatch')
class DetallePedidoUpdateView(LoginRequiredMixin, UpdateView):
    model = DetallePedido
    fields = ['producto', 'cantidad', 'precio_unitario']
    template_name = 'bodega/detallepedido_form.html'
    success_url = reverse_lazy('pedido_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido'] = self.object.pedido  # Mostrar el pedido relacionado
        return context

@method_decorator(permiso_requerido('eliminar_detallepedido'), name='dispatch')
class DetallePedidoDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        detalle = DetallePedido.objects.get(pk=pk)
        detalle.delete()
        return redirect('pedido_list')  # O redirigir a la vista de edición del pedido

@method_decorator(permiso_requerido('ver_factura'), name='dispatch')
class FacturaListView(LoginRequiredMixin, ListView):
    model = Factura
    template_name = 'bodega/factura_list.html'
    context_object_name = 'facturas'
    paginate_by = 10

    def get_queryset(self):
        qs = Factura.objects.select_related('cliente').order_by('-id')
        buscar = self.request.GET.get('buscar')
        if buscar:
            qs = qs.filter(cliente__nombre__icontains=buscar) | qs.filter(estado__icontains=buscar)
        return qs

@method_decorator(permiso_requerido('crear_factura'), name='dispatch')
class FacturaCreateView(LoginRequiredMixin, View):
    def get(self, request):
        form = FacturaForm()
        formset = DetalleFacturaFormSet()
        return render(request, 'bodega/factura_form.html', {'form': form, 'formset': formset})

    def post(self, request):
        form = FacturaForm(request.POST)
        formset = DetalleFacturaFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            factura = form.save()
            formset.instance = factura  # CORRECCIÓN
            formset.save()

            # Calcular total
            total = sum([d.cantidad * d.precio_unitario for d in factura.detalles.all()])
            factura.total = total
            factura.save()

            return redirect('factura_list')

        return render(request, 'bodega/factura_form.html', {'form': form, 'formset': formset})

@method_decorator(permiso_requerido('editar_factura'), name='dispatch')
class FacturaUpdateView(LoginRequiredMixin, View):
    def get(self, request, pk):
        factura = get_object_or_404(Factura, pk=pk)
        form = FacturaForm(instance=factura)
        formset = DetalleFacturaFormSet(instance=factura)
        return render(request, 'bodega/factura_form.html', {'form': form, 'formset': formset})

    def post(self, request, pk):
        factura = get_object_or_404(Factura, pk=pk)
        form = FacturaForm(request.POST, instance=factura)
        formset = DetalleFacturaFormSet(request.POST, instance=factura)

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.instance = factura  # CORRECCIÓN
            formset.save()

            # Recalcular total
            total = sum([d.cantidad * d.precio_unitario for d in factura.detalles.all()])
            factura.total = total
            factura.save()

            return redirect('factura_list')

        return render(request, 'bodega/factura_form.html', {'form': form, 'formset': formset})

@method_decorator(permiso_requerido('cancelar_factura'), name='dispatch')
class FacturaInactivateView(LoginRequiredMixin, View):
    def get(self, request, pk):
        factura = get_object_or_404(Factura, pk=pk)
        return render(request, 'bodega/factura_confirm_inactivate.html', {'factura': factura})

    def post(self, request, pk):
        factura = get_object_or_404(Factura, pk=pk)
        factura.estado = 'anulada' if factura.estado != 'anulada' else 'pendiente'
        factura.save()
        return redirect('factura_list')

@method_decorator(permiso_requerido('ver_inventario'), name='dispatch')
class ReporteInventarioView(LoginRequiredMixin, ListView):
    model = Inventario
    template_name = 'bodega/reportes/reporte_inventario.html'
    context_object_name = 'inventarios'
    paginate_by = 20

    def get_queryset(self):
        # queryset = Inventario.objects.select_related('producto', 'almacen')
        queryset = Inventario.objects.select_related('producto', 'almacen')
        buscar = self.request.GET.get('buscar')
        if buscar:
            queryset = queryset.filter(
                producto__nombre__icontains=buscar
            ) | queryset.filter(
                almacen__nombre__icontains=buscar
            )
        return queryset.order_by('almacen__nombre', 'producto__nombre')

@method_decorator(permiso_requerido('ver_movimiento'), name='dispatch')
class MovimientoReporteView(LoginRequiredMixin, ListView):
    model = Movimiento
    template_name = 'bodega/reportes/reporte_movimientos.html'
    context_object_name = 'movimientos'
    paginate_by = 100

    def get_queryset(self):
        queryset = Movimiento.objects.select_related('producto', 'almacen', 'realizado_por').order_by('-fecha')

        producto = self.request.GET.get('producto')
        tipo = self.request.GET.get('tipo')
        almacen = self.request.GET.get('almacen')
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')

        if producto:
            queryset = queryset.filter(producto__id=producto)
        if tipo:
            queryset = queryset.filter(tipo=tipo)
        if almacen:
            queryset = queryset.filter(almacen__id=almacen)
        if fecha_inicio:
            queryset = queryset.filter(fecha__date__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(fecha__date__lte=fecha_fin)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import Producto, Almacen
        context['productos'] = Producto.objects.all()
        context['almacenes'] = Almacen.objects.all()
        return context

# Vista para el gráfico de Entradas vs Salidas
@login_required
def datos_entradas_salidas(request):
    movimientos = Movimiento.objects.all()

    producto = request.GET.get("producto")
    tipo = request.GET.get("tipo")
    almacen = request.GET.get("almacen")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    if producto:
        movimientos = movimientos.filter(producto_id=producto)
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
    if almacen:
        movimientos = movimientos.filter(almacen_id=almacen)
    if fecha_inicio:
        movimientos = movimientos.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        movimientos = movimientos.filter(fecha__date__lte=fecha_fin)

    entradas = movimientos.filter(tipo='entrada').count()
    salidas = movimientos.filter(tipo='salida').count()

    return JsonResponse({
        'labels': ['Entradas', 'Salidas'],
        'data': [entradas, salidas]
    })

@method_decorator(permiso_requerido('ver_movimiento'), name='dispatch')
class MovimientoReporteExportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimientos"

        # Encabezados
        ws.append(["Fecha", "Producto", "Tipo", "Cantidad", "Almacén", "Observación", "Realizado por"])

        # Datos
        movimientos = Movimiento.objects.select_related('producto', 'almacen', 'realizado_por').all()
        for m in movimientos:
            ws.append([
                localtime(m.fecha).strftime("%Y-%m-%d %H:%M"),
                m.producto.nombre,
                m.tipo.capitalize(),
                m.cantidad,
                m.almacen.nombre,
                m.observacion or '',
                m.realizado_por.username if m.realizado_por else ''
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=movimientos.xlsx'
        wb.save(response)
        return response

@method_decorator(permiso_requerido('ver_inventario'), name='dispatch')
class ReporteInventarioExportView(LoginRequiredMixin, View):
    def get(self, request):
        buscar = request.GET.get("buscar")
        inventarios = Inventario.objects.select_related('producto', 'producto__categoria', 'almacen')

        if buscar:
            inventarios = inventarios.filter(
                models.Q(producto__nombre__icontains=buscar) |
                models.Q(almacen__nombre__icontains=buscar)
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        ws.append(["Producto", "Categoría", "Almacén", "Stock"])

        for inv in inventarios.order_by('producto__nombre'):
            ws.append([
                inv.producto.nombre,
                inv.producto.categoria.nombre if inv.producto.categoria else "",
                inv.almacen.nombre,
                inv.stock
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename={smart_str("reporte_inventario.xlsx")}'
        wb.save(response)
        return response

@method_decorator(permiso_requerido('ver_pedido'), name='dispatch')
class ReportePedidoView(LoginRequiredMixin, ListView):
    model = Pedido
    template_name = 'bodega/reportes/reporte_pedidos.html'
    context_object_name = 'pedidos'
    paginate_by = 20

    def get_queryset(self):
        return Pedido.objects.select_related('cliente').order_by('-fecha')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pedidos = context['pedidos']

        for pedido in pedidos:
            pedido.total_estimado = sum(
                d.cantidad * d.precio_unitario for d in pedido.detalles.all()
            )

        return context

@method_decorator(permiso_requerido('ver_pedido'), name='dispatch')
class ReportePedidoExportView(LoginRequiredMixin, View):
    def get(self, request):
        pedidos = Pedido.objects.select_related('cliente').order_by('-fecha')

        wb = Workbook()
        ws = wb.active
        ws.title = "Pedidos"

        ws.append(["Cliente", "Fecha", "Estado", "Observación", "Total estimado"])

        for pedido in pedidos:
            total_estimado = sum([d.cantidad * d.precio_unitario for d in pedido.detalles.all()])
            ws.append([
                f"{pedido.cliente.nombre} {pedido.cliente.apellido}",
                localtime(pedido.fecha).strftime("%Y-%m-%d %H:%M"),
                pedido.estado.title(),
                pedido.observacion or '',
                total_estimado
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=reporte_pedidos.xlsx'
        wb.save(response)
        return response

@method_decorator(permiso_requerido('ver_factura'), name='dispatch')
class ReporteFacturaView(LoginRequiredMixin, ListView):
    model = Factura
    template_name = 'bodega/reportes/reporte_facturas.html'
    context_object_name = 'facturas'
    paginate_by = 50

    def get_queryset(self):
        return Factura.objects.select_related('cliente').order_by('-fecha')

@method_decorator(permiso_requerido('ver_factura'), name='dispatch')
class ReporteFacturaExportView(LoginRequiredMixin, View):
    def get(self, request):
        facturas = Factura.objects.select_related('cliente').order_by('-fecha')

        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas"

        ws.append(["Cliente", "Fecha", "Estado", "Total"])

        for f in facturas:
            ws.append([
                f"{f.cliente.nombre} {f.cliente.apellido}",
                localtime(f.fecha).strftime("%Y-%m-%d %H:%M"),
                f.estado.title(),
                f.total
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=facturas.xlsx'
        wb.save(response)
        return response

@method_decorator(permiso_requerido('ver_orden_compra'), name='dispatch')
class ReporteOrdenCompraView(LoginRequiredMixin, ListView):
    model = OrdenCompra
    template_name = 'bodega/reportes/reporte_orden_compra.html'
    context_object_name = 'ordenes'
    paginate_by = 50

    def get_queryset(self):
        return OrdenCompra.objects.select_related('proveedor').order_by('-fecha')

@method_decorator(permiso_requerido('ver_orden_compra'), name='dispatch')
class ReporteOrdenCompraExportView(LoginRequiredMixin, View):
    def get(self, request):
        ordenes = OrdenCompra.objects.select_related('proveedor').all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Órdenes de Compra"

        ws.append(["Proveedor", "Fecha", "Estado", "Observación"])

        for o in ordenes:
            ws.append([
                o.proveedor.nombre,
                localtime(o.fecha).strftime("%Y-%m-%d %H:%M"),
                o.estado.title(),
                o.observacion or ''
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=ordenes_compra.xlsx'
        wb.save(response)
        return response

@method_decorator(permiso_requerido('ver_cliente'), name='dispatch')
class ReporteClienteView(LoginRequiredMixin, ListView):
    model = Cliente
    template_name = 'bodega/reportes/reporte_clientes.html'
    context_object_name = 'clientes'
    paginate_by = 50

    def get_queryset(self):
        return Cliente.objects.all().order_by('nombre', 'apellido')

@method_decorator(permiso_requerido('ver_cliente'), name='dispatch')
class ReporteClienteExportView(LoginRequiredMixin, View):
    def get(self, request):
        clientes = Cliente.objects.all().order_by('nombre', 'apellido')

        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        ws.append(["Nombre", "Apellido", "Correo", "Teléfono", "Estado"])

        for c in clientes:
            ws.append([
                c.nombre,
                c.apellido,
                c.email,
                c.telefono,
                "Activo" if c.activo else "Inactivo"
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=clientes.xlsx'
        wb.save(response)
        return response

class ImprimirFacturaView(LoginRequiredMixin, View):
    def get(self, request, pk):
        factura = get_object_or_404(Factura, pk=pk)
        return render(request, 'bodega/factura_print.html', {'factura': factura})

from django.contrib import messages

@method_decorator(permiso_requerido('crear_factura'), name='dispatch')
class GenerarFacturaDesdePedidoView(LoginRequiredMixin, View):
    def get(self, request, pedido_id):
        pedido = get_object_or_404(Pedido, pk=pedido_id)

        # Solo si está entregado
        if pedido.estado != 'entregado':
            messages.warning(request, "Solo se pueden facturar pedidos entregados.")
            return redirect('pedido_list')

        # Crear la factura
        factura = Factura.objects.create(
            cliente=pedido.cliente,
            observacion=pedido.observacion,
            nro_factura='001-001-' + str(Factura.objects.count() + 1).zfill(7),  # Ejemplo básico
            timbrado='12345678',  # Reemplazar por valor real o configurable
            condicion_venta='contado',
            estado='pendiente'
        )

        total = 0
        for detalle in pedido.detalles.all():
            DetalleFactura.objects.create(
                factura=factura,
                producto=detalle.producto,
                cantidad=detalle.cantidad,
                precio_unitario=detalle.precio_unitario,
                iva_aplicado=detalle.producto.iva
            )
            total += detalle.cantidad * detalle.precio_unitario

        factura.total = total
        pedido.facturado = True
        factura.save()

        messages.success(request, "Factura generada correctamente.")
        return redirect('factura_list')

@method_decorator(permiso_requerido('ver_factura'), name='dispatch')
class FacturaPrintView(LoginRequiredMixin, View):
    def get(self, request, pk):
        factura = get_object_or_404(Factura, pk=pk)
        return render(request, 'bodega/factura_print.html', {'factura': factura})

@method_decorator(permiso_requerido('ver_factura'), name='dispatch')
class FacturaPDFView(LoginRequiredMixin, View):
    def get(self, request, pk):
        factura = get_object_or_404(Factura, pk=pk)
        
        # Convertir total a letras en guaraníes
        total_letras = num2words(factura.total, lang='es').replace("coma cero cero", "") + " guaraníes"
        
        template = get_template('bodega/factura_pdf.html')
        html = template.render({
            'factura': factura,
            'total_letras': total_letras
        })

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=factura_{factura.pk}.pdf'

        pisa_status = pisa.CreatePDF(html, dest=response)

        if pisa_status.err:
            return HttpResponse('Ocurrió un error al generar el PDF', status=500)

        return response

@method_decorator(permiso_requerido('abrir_caja'), name='dispatch')
class CajaCreateView(LoginRequiredMixin, CreateView):
    model = Caja
    form_class = CajaForm
    template_name = 'bodega/caja_form.html'

    def form_valid(self, form):
        form.instance.usuario_apertura = self.request.user
        form.instance.fecha_apertura = timezone.now()
        form.instance.estado = 'abierta'
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('caja_list')

@method_decorator(permiso_requerido('cerrar_caja'), name='dispatch')
class CajaCloseView(LoginRequiredMixin, View):
    def post(self, request, pk):
        caja = get_object_or_404(Caja, pk=pk, estado='abierta')
        caja.fecha_cierre = timezone.now()
        caja.usuario_cierre = request.user
        caja.estado = 'cerrada'
        caja.save()
        return redirect('caja_list')

@method_decorator(permiso_requerido('registrar_movimiento_caja'), name='dispatch')
class MovimientoCajaCreateView(LoginRequiredMixin, CreateView):
    model = MovimientoCaja
    form_class = MovimientoCajaForm
    template_name = 'bodega/movimientocaja_form.html'

    def form_valid(self, form):
        form.instance.usuario = self.request.user
        form.instance.fecha = timezone.now()
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('caja_list')

@method_decorator(permiso_requerido('ver_caja'), name='dispatch')
class CajaListView(LoginRequiredMixin, ListView):
    model = Caja
    template_name = 'bodega/caja_list.html'
    context_object_name = 'cajas'

@method_decorator(permiso_requerido('ver_movimientos_caja'), name='dispatch')
class CajaMovimientosView(LoginRequiredMixin, View):
    def get(self, request, pk):
        caja = get_object_or_404(Caja, pk=pk)
        movimientos = caja.movimientos.all()
        return render(request, 'bodega/caja_movimientos.html', {
            'caja': caja,
            'movimientos': movimientos
        })

@method_decorator([login_required, permiso_requerido('ver_tipoproducto')], name='dispatch')
class TipoProductoListView(LoginRequiredMixin, ListView):
    model = TipoProducto
    template_name = 'bodega/tipoproducto_list.html'
    context_object_name = 'tipos'
    paginate_by = 10

    def get_queryset(self):
        queryset = TipoProducto.objects.all()
        buscar = self.request.GET.get('buscar')
        if buscar:
            queryset = queryset.filter(nombre__icontains=buscar)
        return queryset

@method_decorator([login_required, permiso_requerido('crear_tipoproducto')], name='dispatch')
class TipoProductoCreateView(LoginRequiredMixin, CreateView):
    model = TipoProducto
    form_class = TipoProductoForm
    template_name = 'bodega/tipoproducto_form.html'
    success_url = reverse_lazy('tipoproducto_list')

@method_decorator([login_required, permiso_requerido('editar_tipoproducto')], name='dispatch')
class TipoProductoUpdateView(LoginRequiredMixin, UpdateView):
    model = TipoProducto
    form_class = TipoProductoForm
    template_name = 'bodega/tipoproducto_form.html'
    success_url = reverse_lazy('tipoproducto_list')

@method_decorator([login_required, permiso_requerido('inactivar_tipoproducto')], name='dispatch')
class TipoProductoInactivateView(LoginRequiredMixin, View):
    def get(self, request, pk):
        tipo = get_object_or_404(TipoProducto, pk=pk)
        return render(request, 'bodega/tipoproducto_confirm_inactivate.html', {'object': tipo})

    def post(self, request, pk):
        tipo = get_object_or_404(TipoProducto, pk=pk)
        tipo.activo = not tipo.activo
        tipo.save()
        return redirect('tipoproducto_list')
